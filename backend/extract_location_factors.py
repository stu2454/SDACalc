"""
Extract ACTUAL location factors from NDIA SDA Excel calculator
This replaces the sample factors with real NDIA data
"""
import sys
sys.path.insert(0, '/app')

import openpyxl
from decimal import Decimal
from datetime import date
from database import SessionLocal
from models import LocationFactor, SA4Region

def extract_location_factors(excel_path):
    """Extract actual location factors from Excel"""
    
    print("=" * 70)
    print("EXTRACTING ACTUAL LOCATION FACTORS FROM EXCEL")
    print("=" * 70)
    
    # Open Excel file
    print(f"\nOpening Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Get both location factor sheets
    sheets = {}
    if 'Location Factors - New Builds' in wb.sheetnames:
        sheets['NEW_BUILDS'] = wb['Location Factors - New Builds']
    if 'Location Factors - Other' in wb.sheetnames:
        sheets['OTHER'] = wb['Location Factors - Other']
    
    if not sheets:
        print("ERROR: Could not find Location Factors sheets")
        print("Available sheets:", wb.sheetnames)
        return
    
    print(f"✓ Found {len(sheets)} location factor sheets")
    
    # Connect to database
    session = SessionLocal()
    
    try:
        # Delete existing location factors
        print("\nDeleting existing location factors...")
        deleted = session.query(LocationFactor).delete()
        session.commit()
        print(f"✓ Deleted {deleted} existing records")
        
        total_added = 0
        
        for stock_category, sheet in sheets.items():
            print(f"\nProcessing {stock_category} sheet...")
            
            # Extract region names from column B (rows 6-94)
            # Extract factors from columns C-M (building type columns 1-11)
            
            for row_idx in range(6, 95):  # Rows 6-94
                region_name = sheet.cell(row=row_idx, column=2).value  # Column B
                
                if not region_name or not isinstance(region_name, str):
                    continue
                
                region_name = region_name.strip()
                
                # Check if region exists in database
                region = session.query(SA4Region).filter(
                    SA4Region.name == region_name
                ).first()
                
                if not region:
                    print(f"  ⚠️  Region not in database: {region_name}")
                    continue
                
                # Extract factors for each building type column (C=1 through M=11)
                for col_idx, excel_col in enumerate(range(3, 14), start=1):  # Excel columns C-M become building types 1-11
                    factor_value = sheet.cell(row=row_idx, column=excel_col).value
                    
                    if factor_value is None:
                        continue
                    
                    # Convert to Decimal
                    try:
                        if isinstance(factor_value, (int, float)):
                            factor = Decimal(str(factor_value))
                        else:
                            factor = Decimal(factor_value)
                    except:
                        print(f"  ⚠️  Invalid factor for {region_name}, col {excel_col}: {factor_value}")
                        continue
                    
                    # Create location factor record
                    # col_idx is 1-11 (building type column), excel_col is 3-13 (Excel column)
                    lf = LocationFactor(
                        sa4_region=region_name,
                        stock_category=stock_category,
                        building_type_column=col_idx,  # Use 1-11, not Excel column numbers
                        location_factor=factor,
                        effective_from=date(2025, 7, 1),
                        effective_to=None
                    )
                    session.add(lf)
                    total_added += 1
            
            session.commit()
            print(f"  ✓ Added factors for {stock_category}")
        
        print(f"\n✓ Total location factors created: {total_added}")
        
        # Verify some samples
        print("\nSample location factors:")
        samples = session.query(LocationFactor).limit(5).all()
        for lf in samples:
            print(f"  {lf.sa4_region} - {lf.stock_category} - Column {lf.building_type_column}: {lf.location_factor}")
        
        # Check Sydney - Ryde specifically
        print("\nChecking NSW - Sydney - Ryde factors:")
        ryde_factors = session.query(LocationFactor).filter(
            LocationFactor.sa4_region == 'NSW - Sydney - Ryde'
        ).all()
        
        if ryde_factors:
            print(f"  Found {len(ryde_factors)} factors for Ryde")
            for lf in ryde_factors[:3]:
                print(f"  - {lf.stock_category}, Column {lf.building_type_column}: {lf.location_factor}")
        else:
            print("  ⚠️  No factors found for Ryde")
        
        print("\n" + "=" * 70)
        print("LOCATION FACTORS EXTRACTED SUCCESSFULLY")
        print("=" * 70)
        print("\nNow test your calculation - it should match Excel exactly!")
        
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        session.rollback()
        raise
    finally:
        session.close()

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = '/data/SDA_Price_Calculator.xlsx'
    
    extract_location_factors(excel_path)
