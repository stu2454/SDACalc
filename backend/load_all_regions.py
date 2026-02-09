"""
Extract all SA4 regions from NDIA SDA Excel calculator
and load them into the database
"""
import sys
sys.path.insert(0, '/app')

import openpyxl
from database import SessionLocal
from models import SA4Region
from sqlalchemy import text

def extract_and_load_regions(excel_path):
    """Extract SA4 regions from Excel and load into database"""
    
    print("=" * 70)
    print("LOADING ALL SA4 REGIONS FROM EXCEL")
    print("=" * 70)
    
    # Open Excel file
    print(f"\nOpening Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Get Location Factors sheet
    if 'Location Factors - New Builds' in wb.sheetnames:
        sheet = wb['Location Factors - New Builds']
    elif 'Location Factors' in wb.sheetnames:
        sheet = wb['Location Factors']
    else:
        print("ERROR: Could not find Location Factors sheet")
        print("Available sheets:", wb.sheetnames)
        return
    
    print(f"✓ Found sheet: {sheet.title}")
    
    # Extract regions from column B (SA4 Region Names)
    # Regions start at row 6 and go to row 94 (89 regions total)
    regions = []
    print("\nExtracting regions from column B...")
    
    for row in range(6, 95):  # Rows 6-94
        region_name = sheet.cell(row=row, column=2).value  # Column B
        if region_name and isinstance(region_name, str) and region_name.strip():
            region_name = region_name.strip()
            
            # Determine state from region name
            if region_name.startswith('NSW'):
                state = 'NSW'
            elif region_name.startswith('VIC'):
                state = 'VIC'
            elif region_name.startswith('QLD'):
                state = 'QLD'
            elif region_name.startswith('SA'):
                state = 'SA'
            elif region_name.startswith('WA'):
                state = 'WA'
            elif region_name.startswith('TAS'):
                state = 'TAS'
            elif region_name.startswith('NT'):
                state = 'NT'
            elif region_name.startswith('ACT'):
                state = 'ACT'
            else:
                state = 'OTHER'
            
            regions.append({
                'name': region_name,
                'state': state,
                'display_order': len(regions) + 1
            })
    
    print(f"✓ Extracted {len(regions)} regions")
    
    # Show sample
    print("\nSample regions:")
    for region in regions[:5]:
        print(f"  - {region['name']} ({region['state']})")
    print(f"  ... and {len(regions) - 5} more")
    
    # Connect to database
    print("\nConnecting to database...")
    session = SessionLocal()
    
    try:
        # Get existing regions
        existing = {r.name for r in session.query(SA4Region).all()}
        print(f"✓ Found {len(existing)} existing regions")
        
        # Add new regions
        added = 0
        for region_data in regions:
            if region_data['name'] not in existing:
                region = SA4Region(
                    name=region_data['name'],
                    state=region_data['state'],
                    display_order=region_data['display_order']
                )
                session.add(region)
                added += 1
        
        session.commit()
        print(f"✓ Added {added} new regions")
        
        # Show final count
        total = session.query(SA4Region).count()
        print(f"\n✓ Total regions in database: {total}")
        
        # Show by state
        print("\nBreakdown by state:")
        for state in ['NSW', 'VIC', 'QLD', 'SA', 'WA', 'TAS', 'NT', 'ACT']:
            count = session.query(SA4Region).filter(SA4Region.state == state).count()
            if count > 0:
                print(f"  {state}: {count} regions")
        
        print("\n" + "=" * 70)
        print("REGIONS LOADED SUCCESSFULLY")
        print("=" * 70)
        print("\nNext step: Run fix_location_factors.py to generate factors for all regions")
        
    except Exception as e:
        print(f"\nError: {e}")
        session.rollback()
        raise
    finally:
        session.close()

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        # Try common locations
        import os
        possible_paths = [
            '/data/SDA_Price_Calculator.xlsx',
            '/data/sda_calculator.xlsx',
            '/data/SDA Price Calculator.xlsx',
        ]
        
        excel_path = None
        for path in possible_paths:
            if os.path.exists(path):
                excel_path = path
                break
        
        if not excel_path:
            print("ERROR: Excel file not found")
            print("\nUsage: python load_all_regions.py <path_to_excel_file>")
            print("\nOr copy the Excel file to the data/ directory as:")
            print("  - SDA_Price_Calculator.xlsx")
            sys.exit(1)
    
    extract_and_load_regions(excel_path)
